import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import os
from datetime import datetime

# 외부 라이브러리 로딩 (openpyxl)
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None
    print("경고: openpyxl 라이브러리가 필요합니다.")

# [여기 아래에 원본의 class CimonReportMaker: 전체 코드를 붙여넣으세요]
class CimonReportMaker:
    def __init__(self, parent):
        self.parent = parent
        
        paned = ttk.PanedWindow(parent, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=5, pady=5)
        
        # [왼쪽] 입력 패널
        left_frame = ttk.Frame(paned)
        paned.add(left_frame, weight=3)
        
        tk.Label(left_frame, text="태그 입력 (엑셀 컬럼용, 한 줄에 하나씩)", font=("Arial", 10, "bold")).pack(pady=5, anchor="w", padx=10)
        
        text_frame = tk.Frame(left_frame)
        text_frame.pack(fill="both", expand=True, padx=10)
        self.txt_tags = scrolledtext.ScrolledText(text_frame, font=("Arial", 10))
        self.txt_tags.pack(fill="both", expand=True)

        # 생성 방식 선택
        option_frame = tk.Frame(left_frame)
        option_frame.pack(pady=5, fill="x", padx=10)
        
        tk.Label(option_frame, text="생성 방식:", font=("맑은 고딕", 9, "bold")).pack(side="left")
        
        self.gen_mode_var = tk.StringVar(value="SheetAdd")
        ttk.Radiobutton(option_frame, text="시트추가방식 (기존)", variable=self.gen_mode_var, value="SheetAdd").pack(side="left", padx=10)
        ttk.Radiobutton(option_frame, text="새로방식 (아래로 추가)", variable=self.gen_mode_var, value="NewAppend").pack(side="left", padx=10)

        btn_frame = tk.Frame(left_frame); btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="일보 생성", width=12, bg="#e1f5fe", command=lambda: self.gen("Daily")).grid(row=0, column=0, padx=5)
        tk.Button(btn_frame, text="월보 생성", width=12, bg="#e8f5e9", command=lambda: self.gen("Monthly")).grid(row=0, column=1, padx=5)
        tk.Button(btn_frame, text="연보 생성", width=12, bg="#fff3e0", command=lambda: self.gen("Yearly")).grid(row=0, column=2, padx=5)

        # [오른쪽] 설명 패널
        right_frame = ttk.LabelFrame(paned, text="작성 요령", padding="10")
        paned.add(right_frame, weight=2)
        
        desc = (
            "1. Cimon SCADA 의 HMI 태그명을\n"
            "   왼쪽 창에 한 줄에 하나씩 입력하세요.\n\n"
            "2. 엑셀창에서 Ctrl+C를 해서\n"
            "   왼쪽 창 붙여넣기를 추천합니다.\n\n"
            "3. [생성 방식 선택]\n"
            "   - 시트추가: 10개 태그마다 시트 분할\n"
            "   - 새로방식: 한 시트에 아래로 이어붙임\n\n"
            "4. 버튼 클릭 시 해당 포맷 생성\n\n"
            "5. 파일은 프로그램 실행 경로에 저장\n\n"
            "6. 'Report'와 'Cmd' 시트가 생성됩니다."
        )
        ttk.Label(right_frame, text=desc, font=("맑은 고딕", 10), justify="left").pack(anchor="nw")

    def get_tags(self):
        content = self.txt_tags.get("1.0", tk.END)
        tags = [line.strip() for line in content.splitlines() if line.strip()]
        return tags

    def style_range(self, ws, cell_range, border=True, align_center=True, bg_color=None):
        thin = Side(border_style="thin", color="000000")
        border_obj = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        for row in ws[cell_range]:
            for cell in row:
                if border:
                    cell.border = border_obj
                if align_center:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if bg_color:
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

    def gen(self, report_type):
        if not openpyxl:
            messagebox.showerror("오류", "openpyxl 라이브러리가 설치되어 있지 않습니다.\n(pip install openpyxl)")
            return

        tags = self.get_tags()
        if not tags:
            messagebox.showwarning("경고", "태그를 입력해주세요!")
            return

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        chunk_size = 10
        tag_chunks = [tags[i:i + chunk_size] for i in range(0, len(tags), chunk_size)]
        
        mode = self.gen_mode_var.get()

        try:
            if mode == "SheetAdd":
                for idx, chunk in enumerate(tag_chunks):
                    sheet_num = idx + 1
                    ws_r = wb.create_sheet(f"{report_type}_Report_{sheet_num}")
                    self.create_report_sheet(ws_r, chunk, report_type)

                    ws_c = wb.create_sheet(f"{report_type}_Cmd_{sheet_num}")
                    self.create_command_sheet(ws_c, chunk, report_type, report_data_start_row=5)
            
            else: # mode == "NewAppend"
                ws_r = wb.create_sheet(f"{report_type}_Report")
                ws_c = wb.create_sheet(f"{report_type}_Cmd")
                
                rows_count = 0
                if report_type == "Daily": rows_count = 24
                elif report_type == "Monthly": rows_count = 31
                elif report_type == "Yearly": rows_count = 12
                
                block_height = rows_count + 4 + 2
                
                curr_r_row = 4
                curr_c_row = 1
                
                for chunk in tag_chunks:
                    self.create_report_sheet(ws_r, chunk, report_type, start_row_idx=curr_r_row)
                    self.create_command_sheet(ws_c, chunk, report_type, start_row_idx=curr_c_row, report_data_start_row=curr_r_row+1)
                    
                    curr_r_row += block_height
                    curr_c_row += (rows_count * len(chunk))

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            fname = f"용승이가 만든_리포트_{report_type}_{timestamp}.xlsx"
            wb.save(fname)
            
            messagebox.showinfo("완료", f"{fname} 파일이 생성되었습니다.\n총 {len(tags)}개 태그")
            try: os.startfile(os.getcwd())
            except: pass

        except Exception as e:
            messagebox.showerror("에러", f"파일 생성 중 오류 발생:\n{str(e)}")

    def create_report_sheet(self, ws, tags, report_type, start_row_idx=4):
        header_row = start_row_idx
        ws[f'A{header_row}'] = "시간/날짜"
        for i, tag in enumerate(tags):
            col_letter = get_column_letter(2 + i)
            cell = ws[f'{col_letter}{header_row}']
            cell.value = tag
            cell.font = Font(bold=True)
            ws.column_dimensions[col_letter].width = 15

        data_start_row = header_row + 1
        rows_count = 0
        
        if report_type == "Daily":
            rows_count = 24
            for i in range(rows_count): ws[f'A{data_start_row + i}'] = f"{i:02d}:00"
        elif report_type == "Monthly":
            rows_count = 31
            for i in range(rows_count): ws[f'A{data_start_row + i}'] = f"{i+1}일"
        elif report_type == "Yearly":
            rows_count = 12
            for i in range(rows_count): ws[f'A{data_start_row + i}'] = f"{i+1}월"

        data_end_row = data_start_row + rows_count - 1
        max_col_letter = get_column_letter(1 + len(tags))
        self.style_range(ws, f"A{header_row}:{max_col_letter}{data_end_row}")
        self.style_range(ws, f"A{header_row}:{max_col_letter}{header_row}", bg_color="D3D3D3")

        stat_start_row = data_end_row + 1
        stats = [("최소값(MIN)", "MIN"), ("최대값(MAX)", "MAX"), ("평균(AVG)", "AVERAGE"), ("합계(SUM)", "SUM")]
        
        for i, (label, func) in enumerate(stats):
            current_row = stat_start_row + i
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].font = Font(bold=True)
            for j in range(len(tags)):
                col = get_column_letter(2 + j)
                ws[f'{col}{current_row}'] = f"={func}({col}{data_start_row}:{col}{data_end_row})"
        
        self.style_range(ws, f"A{stat_start_row}:{max_col_letter}{stat_start_row + 3}")

    def create_command_sheet(self, ws, tags, report_type, start_row_idx=1, report_data_start_row=5):
        rows_count = 0
        if report_type == "Daily": rows_count = 24
        elif report_type == "Monthly": rows_count = 31
        elif report_type == "Yearly": rows_count = 12
        
        current_write_row = start_row_idx
        fill_blue = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        
        # v1에서는 RichText 사용 X (호환성 유지)
        
        for i, tag in enumerate(tags):
            target_col_letter = get_column_letter(2 + i)
            apply_fill = (i % 2 != 0)

            for r in range(rows_count):
                target_row_num = report_data_start_row + r
                coord_str = f"{target_col_letter}{target_row_num}"
                
                time_str = ""
                if report_type == "Daily": time_str = f'-1일{r:02d}시'
                elif report_type == "Monthly": time_str = f'-1월{r+1}일'
                elif report_type == "Yearly": time_str = f'-1년{r+1}월'

                # v1: 단순 문자열 결합
                cmd_val = f'TlogVal("{tag}", "{time_str}", "순간값")'
                
                ws[f'A{current_write_row}'].value = coord_str
                ws[f'B{current_write_row}'].value = cmd_val
                ws[f'C{current_write_row}'].value = "w"
                
                if apply_fill:
                    ws[f'A{current_write_row}'].fill = fill_blue
                    ws[f'B{current_write_row}'].fill = fill_blue
                    ws[f'C{current_write_row}'].fill = fill_blue
                
                current_write_row += 1
                
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 5
