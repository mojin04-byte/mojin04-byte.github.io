import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
from datetime import datetime

# 외부 라이브러리 로딩 (openpyxl 및 RichText)
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock, CellRichText
except ImportError:
    openpyxl = None
    print("경고: openpyxl 라이브러리가 필요합니다.")

# [여기 아래에 원본의 class CimonReportMakerV2: 전체 코드를 붙여넣으세요]
class CimonReportMakerV2:
    def __init__(self, parent):
        self.parent = parent
        
        # 전체 좌우 분할 (ttk.PanedWindow 사용)
        paned = ttk.PanedWindow(parent, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=5, pady=5)
        
        # [왼쪽] 입력 패널 (너비 가중치 3)
        left_frame = ttk.Frame(paned)
        paned.add(left_frame, weight=3)
        
        # 라벨 변경
        tk.Label(left_frame, text="태그 리스트 (엑셀에서 복사하여 붙여넣기 가능: Tag | 설명 | 타입)", font=("Arial", 10, "bold")).pack(pady=5, anchor="w", padx=10)
        
        # [변경] Treeview(표) 생성 (엑셀 시트 느낌)
        tree_frame = tk.Frame(left_frame)
        tree_frame.pack(fill="both", expand=True, padx=10)
        
        # 스크롤바
        scrollbar_y = ttk.Scrollbar(tree_frame)
        scrollbar_y.pack(side="right", fill="y")
        
        # 컬럼 정의: Tag명, Tag설명, 타입(숫자)
        self.tree = ttk.Treeview(tree_frame, columns=("tag", "desc", "type"), show="headings", yscrollcommand=scrollbar_y.set)
        self.tree.pack(fill="both", expand=True)
        
        scrollbar_y.config(command=self.tree.yview)
        
        # 헤더 설정
        self.tree.heading("tag", text="Tag명 (Cimon)")
        self.tree.heading("desc", text="Tag설명 (리포트표시)")
        self.tree.heading("type", text="타입 (0:순간,1:가동,2:평균,3:적산)")
        
        # 컬럼 너비 설정
        self.tree.column("tag", width=150)
        self.tree.column("desc", width=150)
        self.tree.column("type", width=80, anchor="center")
        
        # 키보드 이벤트 바인딩 (Ctrl+V 붙여넣기, Delete 삭제)
        self.tree.bind("<Control-v>", self.paste_from_clipboard)
        self.tree.bind("<Delete>", self.delete_selected)

        # 수동 입력 프레임 (직접 추가 기능)
        input_frame = tk.Frame(left_frame)
        input_frame.pack(fill="x", padx=10, pady=5)
        
        tk.Label(input_frame, text="Tag:").pack(side="left")
        self.entry_tag = tk.Entry(input_frame, width=15)
        self.entry_tag.pack(side="left", padx=2)
        
        tk.Label(input_frame, text="설명:").pack(side="left")
        self.entry_desc = tk.Entry(input_frame, width=15)
        self.entry_desc.pack(side="left", padx=2)
        
        tk.Label(input_frame, text="타입:").pack(side="left")
        self.entry_type = tk.Entry(input_frame, width=5)
        self.entry_type.pack(side="left", padx=2)
        
        tk.Button(input_frame, text="추가", command=self.add_manual_item).pack(side="left", padx=5)
        tk.Button(input_frame, text="전체삭제", command=self.clear_all).pack(side="right", padx=5)

        # 생성 방식 선택 라디오 버튼
        option_frame = tk.Frame(left_frame)
        option_frame.pack(pady=5, fill="x", padx=10)
        
        tk.Label(option_frame, text="생성 방식:", font=("맑은 고딕", 9, "bold")).pack(side="left")
        
        # [변경] 기본값을 세로방식(NewAppend)으로 변경 및 순서/텍스트 수정
        self.gen_mode_var = tk.StringVar(value="NewAppend") 
        ttk.Radiobutton(option_frame, text="세로방식 (아래로 추가)", variable=self.gen_mode_var, value="NewAppend").pack(side="left", padx=10)
        ttk.Radiobutton(option_frame, text="시트추가방식", variable=self.gen_mode_var, value="SheetAdd").pack(side="left", padx=10)

        # 버튼 영역
        btn_frame = tk.Frame(left_frame)
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="일보 생성", width=12, bg="#e1f5fe", command=lambda: self.gen("Daily")).grid(row=0, column=0, padx=5)
        tk.Button(btn_frame, text="월보 생성", width=12, bg="#e8f5e9", command=lambda: self.gen("Monthly")).grid(row=0, column=1, padx=5)
        tk.Button(btn_frame, text="연보 생성", width=12, bg="#fff3e0", command=lambda: self.gen("Yearly")).grid(row=0, column=2, padx=5)

        # [오른쪽] 설명 패널 (너비 가중치 2)
        right_frame = ttk.LabelFrame(paned, text="작성 요령", padding="10")
        paned.add(right_frame, weight=2)
        
        desc = (
            "1. 엑셀 시트 창으로 변경되었습니다.\n"
            "   [Tag명 | 설명 | 타입] 순서입니다.\n\n"
            "2. 엑셀에서 3개 열을 드래그 복사 후\n"
            "   리스트를 클릭하고 Ctrl+V 하세요.\n\n"
            "3. [타입 숫자 코드]\n"
            "   0 : 순간값\n"
            "   1 : 가동시간\n"
            "   2 : 평균값\n"
            "   3 : 적산값\n\n"
            "4. [생성 방식 선택]\n"
            "   - 세로방식: 한 시트에 아래로 이어붙임\n"
            "   - 시트추가: 10개 태그마다 시트 분할\n\n"
            "5. 생성 버튼을 누르면 저장 위치를\n"
            "   선택할 수 있습니다.\n\n"
            "6. '보고서'와 '명령' 시트가 생성됩니다.\n\n"
            "7. 권이가 원하는 방식 추가"
        )
        ttk.Label(right_frame, text=desc, font=("맑은 고딕", 10), justify="left").pack(anchor="nw")

    def add_manual_item(self):
        """수동 입력 추가"""
        t = self.entry_tag.get().strip()
        d = self.entry_desc.get().strip()
        ty = self.entry_type.get().strip()
        if t:
            self.tree.insert("", "end", values=(t, d, ty))
            self.entry_tag.delete(0, tk.END)
            self.entry_desc.delete(0, tk.END)
            self.entry_type.delete(0, tk.END)
            # 다음 입력을 위해 태그창 포커스
            self.entry_tag.focus()

    def paste_from_clipboard(self, event):
        """클립보드 내용을 붙여넣기 (엑셀 호환)"""
        try:
            clipboard_data = self.parent.clipboard_get()
            rows = clipboard_data.split('\n')
            for row in rows:
                if not row.strip(): continue
                cols = row.split('\t')
                # 3개 열 미만이면 빈칸 채움
                while len(cols) < 3:
                    cols.append("")
                # 앞뒤 공백 제거
                cols = [c.strip() for c in cols]
                self.tree.insert("", "end", values=cols[:3])
        except Exception as e:
            pass # 클립보드가 비었거나 텍스트가 아님

    def delete_selected(self, event):
        """선택 항목 삭제"""
        for item in self.tree.selection():
            self.tree.delete(item)

    def clear_all(self):
        """전체 삭제"""
        for item in self.tree.get_children():
            self.tree.delete(item)

    def get_tag_data(self):
        """Treeview에서 데이터 추출하여 리스트[딕셔너리] 반환"""
        data = []
        for item_id in self.tree.get_children():
            vals = self.tree.item(item_id)['values']
            # values는 튜플로 나옴 (tag, desc, type)
            # 타입이 비어있거나 이상하면 기본값 0 처리
            tag_name = str(vals[0])
            tag_desc = str(vals[1]) if len(vals) > 1 and vals[1] else tag_name
            type_val = str(vals[2]) if len(vals) > 2 else "0"
            
            # 매핑 로직
            suffix = "순간값"
            if type_val == "1": suffix = "가동시간"
            elif type_val == "2": suffix = "평균값"
            elif type_val == "3": suffix = "적산값"
            
            data.append({
                "tag": tag_name,
                "desc": tag_desc,
                "type_code": type_val,
                "suffix": suffix
            })
        return data

    def style_range(self, ws, cell_range, border=True, align_center=True, bg_color=None):
        """엑셀 셀 스타일 적용 헬퍼 함수"""
        thin = Side(border_style="thin", color="000000")
        border_obj = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        for row in ws[cell_range]:
            for cell in row:
                if border:
                    cell.border = border_obj
                if align_center:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if bg_color:
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

    def create_report_sheet(self, ws, tag_items, report_type, start_row_idx=4):
        """보고서(보여지는 화면) 시트 작성"""
        
        # 헤더 행 번호
        header_row = start_row_idx
        
        # 헤더 작성 (A열: 시간, B열~: 태그 설명)
        ws[f'A{header_row}'] = "시간/날짜"
        for i, item in enumerate(tag_items):
            col_letter = get_column_letter(2 + i) # B부터 시작
            cell = ws[f'{col_letter}{header_row}']
            cell.value = item['desc'] # [변경] Tag명 대신 설명(Desc) 사용
            cell.font = Font(bold=True)
            ws.column_dimensions[col_letter].width = 15 # 컬럼 너비 조정

        # 데이터 행 생성 (헤더 다음 행부터 시작)
        data_start_row = header_row + 1
        rows_count = 0
        
        if report_type == "Daily":
            rows_count = 24
            for i in range(rows_count):
                ws[f'A{data_start_row + i}'] = f"{i:02d}:00" # 00:00 ~ 23:00
        
        elif report_type == "Monthly":
            rows_count = 31
            for i in range(rows_count):
                ws[f'A{data_start_row + i}'] = f"{i+1}일" # 1일 ~ 31일

        elif report_type == "Yearly":
            rows_count = 12
            for i in range(rows_count):
                ws[f'A{data_start_row + i}'] = f"{i+1}월" # 1월 ~ 12월

        data_end_row = data_start_row + rows_count - 1

        # 스타일 적용 (테두리 및 정렬)
        # 전체 데이터 영역 (A ~ K)
        max_col_letter = get_column_letter(1 + len(tag_items)) # 태그 개수만큼
        self.style_range(ws, f"A{header_row}:{max_col_letter}{data_end_row}")
        
        # 헤더 색상
        self.style_range(ws, f"A{header_row}:{max_col_letter}{header_row}", bg_color="D3D3D3")

        # 통계 행 추가 (최소값, 최대값, 평균, 합계)
        stat_start_row = data_end_row + 1
        
        stats = [("최소값(MIN)", "MIN"), ("최대값(MAX)", "MAX"), ("평균(AVG)", "AVERAGE"), ("합계(SUM)", "SUM")]
        
        for i, (label, func) in enumerate(stats):
            current_row = stat_start_row + i
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].font = Font(bold=True)
            
            for j in range(len(tag_items)):
                col = get_column_letter(2 + j)
                # 엑셀 수식 입력
                ws[f'{col}{current_row}'] = f"={func}({col}{data_start_row}:{col}{data_end_row})"
        
        # 통계 영역 스타일 적용
        self.style_range(ws, f"A{stat_start_row}:{max_col_letter}{stat_start_row + 3}")

    def create_command_sheet(self, ws, tag_items, report_type, start_row_idx=1, report_data_start_row=5):
        """Cimon 명령어가 들어가는 데이터 시트 작성 (매핑 형태)"""
        rows_count = 0
        if report_type == "Daily": rows_count = 24
        elif report_type == "Monthly": rows_count = 31
        elif report_type == "Yearly": rows_count = 12
        
        current_write_row = start_row_idx
        
        # [수정] 모든 리포트 타입에 대해 첫 줄에 날짜 제목 스크립트 추가
        # start_row_idx가 1인 경우(시트의 맨 처음)에만 추가
        if current_write_row == 1:
            # 리포트 타입별 기준 시간 오프셋 설정
            time_offset = "-1일" # 기본값 (Daily)
            if report_type == "Monthly": time_offset = "-1월"
            elif report_type == "Yearly": time_offset = "-1년"

            ws[f'A{current_write_row}'].value = "B3"
            # 오프셋만 변경하여 스크립트 작성
            ws[f'B{current_write_row}'].value = f'ReportTimeStr("{time_offset}", 12) + "    " + ReportTimeStr("{time_offset}", 64)'
            ws[f'C{current_write_row}'].value = "w"
            current_write_row += 1
        
        fill_blue = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        font_red = InlineFont(color="FF0000") 
        
        for i, item in enumerate(tag_items):
            target_col_letter = get_column_letter(2 + i)
            apply_fill = (i % 2 != 0) 

            tag_name = item['tag']
            suffix_str = item['suffix'] # [변경] 타입에 따른 문자열 (순간값, 가동시간 등)

            for r in range(rows_count):
                # Report 시트에서의 실제 타겟 행 번호
                target_row_num = report_data_start_row + r
                
                # 좌표 문자열 (예: B5, B6...)
                coord_str = f"{target_col_letter}{target_row_num}"
                
                time_str = ""
                if report_type == "Daily": time_str = f'-1일{r:02d}시'
                elif report_type == "Monthly": time_str = f'-1월{r+1}일'
                elif report_type == "Yearly": time_str = f'-1년{r+1}월'

                part1 = 'TlogVal("'
                part2 = tag_name 
                part3 = f'", "{time_str}", "{suffix_str}")' # [변경] suffix_str 적용
                
                # Cimon 명령어용 RichText 생성 (태그명만 빨간색)
                rich_string = CellRichText(part1, TextBlock(font_red, part2), part3)
                
                ws[f'A{current_write_row}'].value = coord_str
                ws[f'B{current_write_row}'].value = rich_string 
                ws[f'C{current_write_row}'].value = "w"
                
                if apply_fill:
                    ws[f'A{current_write_row}'].fill = fill_blue
                    ws[f'B{current_write_row}'].fill = fill_blue
                    ws[f'C{current_write_row}'].fill = fill_blue
                
                current_write_row += 1
                
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 5

    def gen(self, report_type):
        if not openpyxl:
            messagebox.showerror("오류", "openpyxl 라이브러리가 설치되어 있지 않습니다.\n(pip install openpyxl)")
            return

        tag_items = self.get_tag_data() # [변경] 데이터 구조체 가져오기
        if not tag_items:
            messagebox.showwarning("경고", "태그를 입력해주세요!")
            return

        # [변경] 한국어 리포트 타입 설정 및 파일명 접두어 추가
        type_kr = "일보"
        if report_type == "Monthly": type_kr = "월보"
        elif report_type == "Yearly": type_kr = "연보"

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"용승이가만든_{type_kr}_{timestamp}.xlsx"
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            initialfile=default_name,
            title="저장할 위치와 파일명을 선택하세요"
        )
        
        if not file_path: # 취소한 경우
            return

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        chunk_size = 10
        tag_chunks = [tag_items[i:i + chunk_size] for i in range(0, len(tag_items), chunk_size)]
        mode = self.gen_mode_var.get()

        try:
            if mode == "SheetAdd":
                # [기존 방식] 10개 태그마다 시트 분리
                for idx, chunk in enumerate(tag_chunks):
                    sheet_num = idx + 1
                    # [변경] 시트 이름 한글화
                    ws_r = wb.create_sheet(f"{type_kr}_보고서_{sheet_num}")
                    self.create_report_sheet(ws_r, chunk, report_type) 
                    
                    ws_c = wb.create_sheet(f"{type_kr}_명령_{sheet_num}")
                    self.create_command_sheet(ws_c, chunk, report_type, start_row_idx=1, report_data_start_row=5)
            else:
                # [새로 방식] 단일 시트에 이어붙이기
                # [변경] 시트 이름 한글화
                ws_r = wb.create_sheet(f"{type_kr}_보고서")
                ws_c = wb.create_sheet(f"{type_kr}_명령")
                
                rows_count = 24 if report_type=="Daily" else 31 if report_type=="Monthly" else 12
                # 데이터(rows_count) + 통계(4) + 여백(2)
                block_height = rows_count + 4 + 2 
                
                curr_r_row = 4  # Report 시트 작성 시작 행
                curr_c_row = 1  # Cmd 시트 작성 시작 행
                
                for chunk in tag_chunks:
                    # 1. Report 시트 작성
                    self.create_report_sheet(ws_r, chunk, report_type, start_row_idx=curr_r_row)
                    
                    # 2. Cmd 시트 작성
                    report_data_start = curr_r_row + 1
                    
                    self.create_command_sheet(ws_c, chunk, report_type, start_row_idx=curr_c_row, report_data_start_row=report_data_start)
                    
                    # 다음 루프를 위해 위치 갱신
                    curr_r_row += block_height
                    # Cmd 시트는 (데이터행 수 * 태그 수) 만큼 아래로 내려감
                    
                    added_rows = rows_count * len(chunk)
                    # [수정] 모든 타입에서 첫 번째 블록(시트 시작)일 경우 헤더 1줄이 추가됨을 반영
                    if curr_c_row == 1:
                         added_rows += 1
                         
                    curr_c_row += added_rows

            wb.save(file_path)
            
            messagebox.showinfo("완료", f"파일 생성 완료!\n경로: {file_path}")
            
            # 생성된 파일이 있는 폴더 열기 (Windows)
            try:
                os.startfile(os.path.dirname(file_path))
            except AttributeError:
                pass 
                
        except Exception as e:
            messagebox.showerror("에러", f"파일 생성 중 오류 발생:\n{str(e)}")
